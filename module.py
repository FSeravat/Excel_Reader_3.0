from unidecode import unidecode
from openpyxl import load_workbook
import json

class Excel_File:

    def __init__(self, file):
        self.dictionary = json.load(open("dictionary.json", encoding='utf-8'))
        self.read_excel = load_workbook(file)
        self.concat = False
        self.findSheets()
        self.findHeader()
        self.formatTable()
        self.readValues()

    def findSheets(self):
        sheet_list = []
        blacklist = ["capa", "plan1", "plan3","a01","fl capa u-8601","capa-5135","n-1710","planilha1","capa ld se"]
        for sheet in self.read_excel._sheets: 
            if sheet.sheet_state == "visible" and not blacklist.__contains__(sheet.title.lower().strip()):
                sheet_list.append(sheet)
        self.sheet_list = sheet_list
        
        
    
    def findHeader(self):
        worksheet = self.sheet_list[0]
        for x in range(worksheet.min_column, worksheet.max_column):
            for y in range(worksheet.min_row, 10):
                cell = worksheet.cell(row=y, column=x)
                value = unidecode(str(cell.value).lower())
                if value.__contains__("titulo"):
                    self.checkMerged(cell)
    
    def checkMerged(self, cell):
        coordinate = cell.coordinate
        self.header = cell.row
        self.titleColumn = cell.column+1
        worksheet = self.sheet_list[0]
        for merged in worksheet.merged_cells:
            if merged.coord.__contains__(coordinate):
                self.header = merged.max_row
                self.titleColumn = merged.max_col+1

    def formatTable(self):
        for sheet in self.sheet_list:
            widthList = [30]
            for width in sheet.column_dimensions:
                widthList.append(sheet.column_dimensions[width].width)
            sheet.insert_cols(1)
            count = 0
            for width in sheet.column_dimensions:
                sheet.column_dimensions[width].width = widthList[count]
                count+=1

    def readValues(self):
        for sheet in self.sheet_list:
            for i in range(self.header+1, sheet.max_row):
                value = sheet.cell(row=i,column=self.titleColumn).value
                findKey = ""
                for key in self.dictionary.keys():
                    check = True
                    for match in self.dictionary[key]:
                        if key == "Plantas, Cortes e Detalhes" or key == "Diagrama de Topologia do Sistema/Rede":
                            check = False
                            if unidecode(str(value).lower()).__contains__(match):
                                findKey = key
                                break
                        else:
                            if not unidecode(str(value).lower()).__contains__(match):
                                check = False
                                break
                    if check == True:
                        findKey = key
                    if not findKey == "":
                        break
                sheet.cell(row=i,column=1).value = findKey